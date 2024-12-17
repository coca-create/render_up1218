import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
import tempfile
import json
from pydub import AudioSegment
import gradio as gr
#from faster_whisper import WhisperModel
import zipfile
from docx import Document
import re
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
from tqdm import tqdm
from datetime import datetime
import traceback
import shutil


def get_audio_duration(filepath):
    try:

        audio = AudioSegment.from_file(filepath)
        duration = len(audio) / 1000.0
        return duration

    except Exception as e:
        print(f"get_audio_durationでerror:{e}")
        traceback.print_exc()
        return None

def format_timestamp(seconds):
    hrs, secs = divmod(seconds, 3600)
    mins, secs = divmod(secs, 60)
    millis = int((secs % 1) * 1000)
    return f"{int(hrs):02}:{int(mins):02}:{int(secs):02},{millis:03}"

#dataframe追加
def parse_srt_c(srt_content):
    pattern = r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)\n\n'
    matches = re.findall(pattern, srt_content, re.DOTALL)
    
    subtitles = []
    for match in matches:
        subtitles.append({
            'ID': int(match[0]),
            'Start': match[1],
            'End': match[2],
            'Text': match[3].replace('\n', ' ')
        })
    
    return subtitles

def dataframe_to_html_table(df):
    return df.to_html(index=False)

# SRTファイルからExcelファイルを作成する関数
def create_excel_from_srt_c(srt_content, input_file_name):
    excel_file_name = f"{input_file_name}_srt.xlsx"
    english_subtitles = parse_srt_c(srt_content)

    data = []
    for eng in english_subtitles:
        data.append({
            'ID': eng['ID'],
            'Start': eng['Start'],
            'End': eng['End'],
            'English Subtitle': eng['Text']
        })

    df = pd.DataFrame(data)
    temp_dir = tempfile.gettempdir()
    excel_file_path = os.path.join(temp_dir, excel_file_name)
    
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Subtitles')
        workbook = writer.book
        worksheet = writer.sheets['Subtitles']

        column_widths = {'A': 7, 'B': 25, 'C': 25, 'D': 90, 'E': 90}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.column_letter in ['B', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column_letter in ['D', 'E']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            worksheet.row_dimensions[row[0].row].height = 30

        header_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    
    return excel_file_path, df

'''def exe_for_gradio(srt_content, input_file_name='Noname'):
    excel_filepath, df_display = create_excel_from_srt_c(srt_content, input_file_name)
    return df_display'''



'''
テキストエリア①
テキストエリア②
テキストエリア③
ファイル④
ファイル⑤
HTML⑥
HTML⑦
HTML⑧
テキストボックス⑨
テキストボックス⑩
HTML⑪
'''
